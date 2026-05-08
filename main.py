# main.py

import os
import sqlite3
from datetime import datetime, timedelta

import discord
from discord.ext import commands, tasks
from discord import app_commands

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# TOKEN
# =========================

TOKEN = os.getenv("DISCORD_TOKEN")

# =========================
# BOT
# =========================

intents = discord.Intents.default()
intents.members = True
intents.message_content = True

bot = commands.Bot(command_prefix="!", intents=intents)

# =========================
# DB
# =========================

conn = sqlite3.connect("poll.db")
c = conn.cursor()

c.execute("""
CREATE TABLE IF NOT EXISTS votes (
    user_id TEXT,
    user_name TEXT,
    day INTEGER,
    time_slot TEXT
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS polls (
    message_id TEXT,
    day INTEGER,
    deadline TEXT,
    closed INTEGER DEFAULT 0
)
""")

conn.commit()

# =========================
# 設定
# =========================

TIME_SLOTS = [
    "午前中",
    "12-14",
    "14-16",
    "16-18",
    "18-20",
    "20〜",
    "21〜",
    "22〜",
    "23〜",
    "24〜"
]

DAY_TITLES = [
    "【1日目㈯】",
    "【2日目㈰】",
    "【3日目㈪】",
    "【4日目㈫】"
]

# =========================
# 締切計算
# =========================

def get_deadlines():

    now = datetime.now()

    monday = now - timedelta(days=now.weekday())

    friday = monday + timedelta(days=4, hours=18)
    saturday = monday + timedelta(days=5, hours=18)
    sunday = monday + timedelta(days=6, hours=18)
    next_monday = monday + timedelta(days=7, hours=18)

    return [
        friday,
        saturday,
        sunday,
        next_monday
    ]

# =========================
# Embed生成
# =========================

def create_poll_embed(day, deadline_text):

    embed = discord.Embed(
        title=f"{DAY_TITLES[day - 1]} {deadline_text}",
        color=0xf4c7ab
    )

    text = ""

    for slot in TIME_SLOTS:

        users = c.execute("""
        SELECT user_name
        FROM votes
        WHERE day=? AND time_slot=?
        ORDER BY user_name
        """, (day, slot)).fetchall()

        users = [u[0] for u in users]

        text += f"### {slot} ({len(users)})\n"

        if users:
            for user in users:
                text += f"・{user}\n"
        else:
            text += "なし\n"

        text += "\n"

    embed.description = text

    embed.set_footer(
        text="ボタンで複数選択可能"
    )

    return embed

# =========================
# ボタン
# =========================

class TimeButton(discord.ui.Button):

    def __init__(self, day, slot):

        super().__init__(
            label=slot,
            style=discord.ButtonStyle.secondary,
            custom_id=f"{day}_{slot}"
        )

        self.day = day
        self.slot = slot

    async def callback(self, interaction: discord.Interaction):

        user_id = str(interaction.user.id)
        user_name = interaction.user.display_name

        c.execute("""
        SELECT *
        FROM votes
        WHERE user_id=?
        AND day=?
        AND time_slot=?
        """, (user_id, self.day, self.slot))

        exists = c.fetchone()

        # =========================
        # 投票解除
        # =========================

        if exists:

            c.execute("""
            DELETE FROM votes
            WHERE user_id=?
            AND day=?
            AND time_slot=?
            """, (user_id, self.day, self.slot))

        # =========================
        # 投票追加
        # =========================

        else:

            c.execute("""
            INSERT INTO votes
            VALUES (?, ?, ?, ?)
            """, (
                user_id,
                user_name,
                self.day,
                self.slot
            ))

        conn.commit()

        # =========================
        # ボタン更新
        # =========================

        for item in self.view.children:

            if isinstance(item, TimeButton):

                users = c.execute("""
                SELECT COUNT(*)
                FROM votes
                WHERE day=?
                AND time_slot=?
                """, (
                    item.day,
                    item.slot
                )).fetchone()[0]

                item.label = f"{item.slot} ({users})"

        # =========================
        # Embed更新
        # =========================

        deadlines = get_deadlines()

        deadline_text = deadlines[self.day - 1].strftime(
            "〆%m/%d 18時"
        )

        embed = create_poll_embed(
            self.day,
            deadline_text
        )

        await interaction.response.edit_message(
            embed=embed,
            view=self.view
        )

# =========================
# View
# =========================

class PollView(discord.ui.View):

    def __init__(self, day):

        super().__init__(timeout=None)

        for slot in TIME_SLOTS:

            users = c.execute("""
            SELECT COUNT(*)
            FROM votes
            WHERE day=?
            AND time_slot=?
            """, (day, slot)).fetchone()[0]

            button = TimeButton(day, slot)

            button.label = f"{slot} ({users})"

            self.add_item(button)

# =========================
# create_poll
# =========================

@bot.tree.command(
    name="create_poll",
    description="4日分アンケート作成"
)
async def create_poll(interaction: discord.Interaction):

    deadlines = get_deadlines()

    for i in range(4):

        deadline_text = deadlines[i].strftime(
            "〆%m/%d 18時"
        )

        embed = create_poll_embed(
            i + 1,
            deadline_text
        )

        view = PollView(i + 1)

        msg = await interaction.channel.send(
            embed=embed,
            view=view
        )

        c.execute("""
        INSERT INTO polls
        (message_id, day, deadline)
        VALUES (?, ?, ?)
        """, (
            str(msg.id),
            i + 1,
            deadlines[i].isoformat()
        ))

        conn.commit()

    await interaction.response.send_message(
        "✅ アンケートを作成しました",
        ephemeral=True
    )

# =========================
# Excel出力
# =========================

@bot.tree.command(
    name="export_excel",
    description="Excel出力"
)
async def export_excel(interaction: discord.Interaction):

    wb = openpyxl.Workbook()

    orange = PatternFill(
        start_color="F4C7AB",
        end_color="F4C7AB",
        fill_type="solid"
    )

    gray = PatternFill(
        start_color="E7E6E6",
        end_color="E7E6E6",
        fill_type="solid"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    users = c.execute("""
    SELECT DISTINCT user_name
    FROM votes
    ORDER BY user_name
    """).fetchall()

    users = [u[0] for u in users]

    for day in range(1, 5):

        if day == 1:
            ws = wb.active
            ws.title = DAY_TITLES[day - 1]
        else:
            ws = wb.create_sheet(
                title=DAY_TITLES[day - 1]
            )

        ws.cell(row=1, column=1, value="名前")

        for i, slot in enumerate(TIME_SLOTS):
            ws.cell(
                row=1,
                column=i + 2,
                value=slot
            )

        for row_idx, user in enumerate(users, start=2):

            name_cell = ws.cell(
                row=row_idx,
                column=1,
                value=user
            )

            user_has_vote = False

            for col_idx, slot in enumerate(
                TIME_SLOTS,
                start=2
            ):

                cell = ws.cell(
                    row=row_idx,
                    column=col_idx
                )

                result = c.execute("""
                SELECT *
                FROM votes
                WHERE user_name=?
                AND day=?
                AND time_slot=?
                """, (
                    user,
                    day,
                    slot
                )).fetchone()

                if result:
                    cell.fill = orange
                    user_has_vote = True
                else:
                    cell.fill = gray

                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center"
                )

            if user_has_vote:
                name_cell.fill = orange

            name_cell.border = border

        for col in range(
            1,
            len(TIME_SLOTS) + 2
        ):
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 14

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center"
                )

        for cell in ws[1]:
            cell.font = Font(bold=True)

    filename = "schedule.xlsx"

    wb.save(filename)

    await interaction.response.send_message(
        file=discord.File(filename)
    )

# =========================
# 締切監視
# =========================

@tasks.loop(minutes=1)
async def check_deadlines():

    now = datetime.now()

    polls = c.execute("""
    SELECT message_id, day, deadline
    FROM polls
    WHERE closed=0
    """).fetchall()

    for message_id, day, deadline in polls:

        deadline_dt = datetime.fromisoformat(deadline)

        if now >= deadline_dt:

            c.execute("""
            UPDATE polls
            SET closed=1
            WHERE message_id=?
            """, (message_id,))

            conn.commit()

            print(f"Day {day} closed")

# =========================
# 起動
# =========================

@bot.event
async def on_ready():

    await bot.tree.sync()

    check_deadlines.start()

    print(f"Logged in as {bot.user}")

# =========================
# RUN
# =========================

bot.run(TOKEN)
